#!/usr/bin/env python3
"""
Mass Recall Exposure Analysis Script

Cross-references a Recall Report against SHD stock data to identify
outlets holding positive SOH of recalled articles.

Usage:
    python process_recall.py [--recall <path>] [--shd <path>] [--output <path>]

If no paths are provided, auto-detects the most recent files in Downloads.
"""

import sys
import os
import argparse
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from pathlib import Path


def find_latest_file(pattern, directory=None):
    """Find the most recently modified file matching a glob pattern."""
    if directory is None:
        directory = r'C:\\Users\\USER\\Downloads'
    p = Path(directory)
    matches = sorted(p.glob(pattern), key=lambda x: x.stat().st_mtime, reverse=True)
    return str(matches[0]) if matches else None


def style_excel(out_path, out_df):
    """Apply styling to the generated Excel file."""
    wb = load_workbook(out_path)
    ws = wb.active

    header_font = Font(name="Arial", size=11, color="FFFFFF", bold=True)
    data_font = Font(name="Arial", size=10)
    header_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            if col_idx <= 2:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            try:
                val = str(cell.value) if cell.value else ''
                max_length = max(max_length, len(val))
            except Exception:
                pass
        adjusted_width = min(max(max_length + 4, 12), 55)
        ws.column_dimensions[chr(64 + col_idx)].width = adjusted_width

    last_col_letter = chr(64 + ws.max_column)
    ws.auto_filter.ref = f'A1:{last_col_letter}{ws.max_row}'

    wb.save(out_path)


def process_recall(recall_path=None, shd_path=None, output_path=None):
    """Main processing function."""

    # Auto-detect files if not provided
    if recall_path is None:
        recall_path = find_latest_file('Recall_Report_*.xlsx')
        if recall_path is None:
            print("ERROR: No Recall_Report_*.xlsx found in Downloads")
            return

    if shd_path is None:
        shd_path = find_latest_file('SHD_*.xlsx')
        if shd_path is None:
            print("ERROR: No SHD_*.xlsx found in Downloads")
            return

    if output_path is None:
        from datetime import datetime
        today = datetime.now().strftime('%d%m%Y')
        output_path = rf'C:\\Users\\USER\\Downloads\\Mass_Recall_Check_{today}.xlsx'

    print(f"Recall file: {recall_path}")
    print(f"SHD file: {shd_path}")
    print(f"Output: {output_path}")
    print()

    # Step 1: Load recall report
    recall = pd.read_excel(recall_path)
    print(f"Recall report loaded: {len(recall)} rows")

    recalled_codes = recall["ArticleCode"].unique().tolist()
    print(f"Unique recalled articles: {len(recalled_codes)} ({recalled_codes})")
    print()

    # Step 2: Load SHD (only needed columns for performance)
    shd = pd.read_excel(
        shd_path,
        usecols=['Store', 'ArticleCode', 'ArticleDesc', 'SOH'],
        dtype={'ArticleCode': str, 'Store': str},
        engine='openpyxl'
    )
    print(f"SHD loaded: {len(shd)} rows")

    # Step 3: Cross-reference - filter for recalled articles with positive SOH
    shd_matched = shd[shd['ArticleCode'].isin([str(c) for c in recalled_codes])].copy()
    shd_matched = shd_matched[shd_matched['SOH'] > 0].copy()

    print(f"Matched rows (SOH > 0): {len(shd_matched)}")
    print()

    if len(shd_matched) == 0:
        print("WARNING: No outlets found with positive SOH for recalled articles.")
        print("This may indicate all recalled items have been cleared from stock.")
        out = pd.DataFrame(columns=recall.columns.tolist())
        out.to_excel(output_path, index=False, sheet_name="Mass Recall Check")
        style_excel(output_path, out)
        return

    # Step 4: Build output rows by merging recall details into SHD data
    rows = []
    for _, r in shd_matched.iterrows():
        code = int(r['ArticleCode'])
        recall_info = recall[recall['ArticleCode'] == code].iloc[0]

        store_str = str(r['Store']).strip()
        if ' - ' in store_str:
            store_num = int(store_str.split(' - ')[0].strip())
        else:
            store_num = int(store_str)

        rows.append({
            'ArticleCode': code,
            'ArticleDesc': r['ArticleDesc'],
            'Return Reason': str(recall_info.get('Return Reason', '')),
            'Buyer Remark / Batch': str(recall_info.get('Buyer Remark / Batch', '')),
            'Outlet Involved': r['Store'],
            'Store Code': store_num,
            'SOH': r['SOH'],
            'Allocation / Recall No': str(recall_info.get('Allocation / Recall No', '')),
            'Store Instruction': str(recall_info.get('Store Instruction', '')),
            'Deadline': recall_info.get('Deadline', ''),
            'Completion Status': '',
            'Remark': ''
        })

    out = pd.DataFrame(rows)

    # Sort by ArticleCode, then Store Code
    out = out.sort_values(['ArticleCode', 'Store Code'], ascending=[True, True]).reset_index(drop=True)

    # Step 5: Save styled Excel
    out.to_excel(output_path, index=False, sheet_name='Mass Recall Check')
    style_excel(output_path, out)

    print(f"Saved: {output_path}")
    print(f'Total rows: {len(out)}, Unique articles: {out["ArticleCode"].nunique()}')
    print()

    # Step 6: Print summary table
    print("=" * 120)
    hdr = f'{"ArticleCode":>10} | {"ArticleDesc":<50} | {"Return Reason":<25} | {"SOH":>6} | {"Store":>6} | {"Deadline"}'
    print(hdr)
    print("=" * 120)
    for _, r in out.iterrows():
        reason = str(r['Return Reason'])[:25]
        deadline = str(r['Deadline']) if pd.notna(r['Deadline']) else ''
        print(f'{int(r["ArticleCode"]):>10} | {str(r["ArticleDesc"])[:50]:<50} | {reason:<25} | {r["SOH"]:>6.2f} | {int(r["Store Code"]):>6} | {deadline}')
    print("=" * 120)

    # Step 7: Summary by article
    print()
    print("SUMMARY BY ARTICLE:")
    print("-" * 80)
    for code in out['ArticleCode'].unique():
        sub = out[out['ArticleCode'] == code]
        desc = sub.iloc[0]['ArticleDesc']
        total_soh = sub['SOH'].sum()
        stores = ', '.join([str(int(s)) for s in sub['Store Code'].unique()])
        reason = sub.iloc[0]['Return Reason']
        print(f'  {int(code)} | {desc[:45]:<45} | SOH: {total_soh:>6.2f} | Stores: {stores}')
    print("-" * 80)


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description="Mass Recall Exposure Analysis")
    parser.add_argument('--recall', type=str, help='Path to recall report Excel file')
    parser.add_argument('--shd', type=str, help='Path to SHD Excel file')
    parser.add_argument('--output', type=str, help='Output Excel file path')
    args = parser.parse_args()

    process_recall(
        recall_path=args.recall,
        shd_path=args.shd,
        output_path=args.output
    )
