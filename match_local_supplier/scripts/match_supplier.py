import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment
import argparse
import os

def main():
    parser = argparse.ArgumentParser(description='Match reordering list with Sabah local ordering POISON sheet.')
    parser.add_argument('reorder_file', type=str, help='Path to the Reorder formatted Excel file.')
    parser.add_argument('sabah_file', type=str, help='Path to the Sabah Local Ordering Excel file.')
    parser.add_argument('--output', type=str, help='Path to the output summary file.', default=None)
    args = parser.parse_args()

    f2 = args.reorder_file
    f1 = args.sabah_file
    
    if args.output:
        out_path = args.output
    else:
        # Default output next to reorder file
        base_dir = os.path.dirname(f2)
        today_str = datetime.now().strftime('%d%m%Y')
        out_path = os.path.join(base_dir, f'Sabah_Local_Ordering_Summary_{today_str}.xlsx')

    # Read files
    print(f"Reading {f2}...")
    df_reorder = pd.read_excel(f2)
    print(f"Reading {f1} sheet 'POISON TO ORDER LOCALLY '...")
    df_sabah = pd.read_excel(f1, sheet_name='POISON TO ORDER LOCALLY ')

    # Filter only for MEDICINE category, ignore OTC MEDICINE
    if 'Category' in df_reorder.columns:
        df_reorder = df_reorder[df_reorder['Category'].str.strip().str.upper() == 'MEDICINE']

    # Clean Sabah columns
    df_sabah = df_sabah.rename(columns={
        'Article Code': 'ArticleCode',
        'SAME COST (Y/N)': 'Same Cost',
        'Nett  cost (YES/NO)': 'Nett Pricing',
        'SUPPLIER': 'Supplier'
    })

    # Clean Article codes for matching
    df_sabah['ArticleCode'] = df_sabah['ArticleCode'].astype(str).str.split('.').str[0].str.strip()
    df_reorder['ArticleCode'] = df_reorder['ArticleCode'].astype(str).str.split('.').str[0].str.strip()

    # Clean Supplier Names
    if 'Supplier' in df_sabah.columns:
        df_sabah['Supplier'] = df_sabah['Supplier'].astype(str).str.strip()

    # Extract relevant mapping data from Sabah sheet
    df_mapping = df_sabah[['ArticleCode', 'Supplier', 'Same Cost', 'Nett Pricing']].dropna(subset=['ArticleCode', 'Supplier']).drop_duplicates(subset=['ArticleCode'])

    # Merge to find matching reorder items (Left join to include all)
    df_matched = pd.merge(df_reorder, df_mapping, on='ArticleCode', how='left')
    df_matched.fillna({'Supplier': 'N/A', 'Same Cost': 'N/A', 'Nett Pricing': 'N/A'}, inplace=True)

    # Calculate 30 Days Top Up Qty
    if 'Daily Demand' in df_matched.columns and 'Pipeline Stock' in df_matched.columns:
        df_matched['Top Up Qty (30 Days)'] = (df_matched['Daily Demand'] * 30) - df_matched['Pipeline Stock']
        df_matched['Top Up Qty (30 Days)'] = df_matched['Top Up Qty (30 Days)'].clip(lower=0).apply(np.ceil).astype(int)
    else:
        print("Warning: Daily Demand or Pipeline Stock not found. Top Up Qty (30 Days) cannot be calculated.")
        df_matched['Top Up Qty (30 Days)'] = 0

    # Ensure 60 days column exists
    if 'Top Up Qty (60 Days)' not in df_matched.columns:
        df_matched['Top Up Qty (60 Days)'] = 0

    # Create a Summary grouped by Store, Supplier and Article
    df_summary = df_matched.groupby(
        ['Store', 'Supplier', 'ArticleCode', 'ArticleDesc', 'Same Cost', 'Nett Pricing'], 
        dropna=False
    ).agg({
        'Top Up Qty (30 Days)': 'sum',
        'Top Up Qty (60 Days)': 'sum'
    }).reset_index()

    # Sort by Store, Supplier and then by Qty
    df_summary = df_summary.sort_values(by=['Store', 'Supplier', 'Top Up Qty (60 Days)'], ascending=[True, True, False])

    # Format the Detailed List
    cols_detailed = [
        'Store', 'ArticleCode', 'ArticleDesc', 'Brand', 'Category', 'SOS', 'SOH',
        'Top Up Qty (30 Days)', 'Top Up Qty (60 Days)', 'Supplier', 'Same Cost', 'Nett Pricing'
    ]
    df_detailed = df_matched[[c for c in cols_detailed if c in df_matched.columns]].sort_values(by=['Supplier', 'Store', 'ArticleCode'])

    # Save to Excel with openpyxl to apply formatting
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = 'Summary By Supplier'

    # Write Summary
    for r in dataframe_to_rows(df_summary, index=False, header=True):
        ws_summary.append(r)

    ws_detailed = wb.create_sheet(title='Detailed List')
    for r in dataframe_to_rows(df_detailed, index=False, header=True):
        ws_detailed.append(r)

    header_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    for ws in [ws_summary, ws_detailed]:
        # Style Header
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
        
        # Auto-filter
        ws.auto_filter.ref = ws.dimensions
        
        # Style Rows and Auto-width
        col_widths = {}
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.value is not None:
                    # Find column name
                    col_name = ws.cell(row=1, column=cell.column).value
                    if col_name == 'ArticleDesc':
                        cell.alignment = left_align
                    else:
                        cell.alignment = center_align
                        
                    # Track width
                    val_len = len(str(cell.value))
                    if col_widths.get(cell.column, 0) < val_len:
                        col_widths[cell.column] = val_len

        # Apply Header widths
        for cell in ws[1]:
            val_len = len(str(cell.value))
            if col_widths.get(cell.column, 0) < val_len:
                col_widths[cell.column] = val_len
                
        for col, width in col_widths.items():
            # Add a little padding
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = min(width + 2, 50) # Cap width at 50

    wb.save(out_path)
    print(f'Report successfully generated at: {out_path}')

    # Output markdown summary
    print('\n=== SUMMARY ===')
    summary_agg = df_summary.groupby('Supplier')[['Top Up Qty (30 Days)', 'Top Up Qty (60 Days)']].sum().reset_index()
    summary_agg = summary_agg.sort_values(by='Top Up Qty (60 Days)', ascending=False)
    print('| Supplier | Total Top Up Qty (30 Days) | Total Top Up Qty (60 Days) |')
    print('| --- | --- | --- |')
    for _, row in summary_agg.iterrows():
        print(f'| {row["Supplier"]} | {row["Top Up Qty (30 Days)"]} | {row["Top Up Qty (60 Days)"]} |')

if __name__ == '__main__':
    main()
